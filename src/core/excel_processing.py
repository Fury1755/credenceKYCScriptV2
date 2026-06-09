"""
This module handles the logic and processing of workbooks.
The input and output should be workbooks.
"""

from openpyxl import Workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.cell.cell import Cell, MergedCell
from openpyxl.utils import range_boundaries
from core.individual import Individual
from typing import Union, Optional, List, TYPE_CHECKING
import logging

if TYPE_CHECKING:
    from boundary.sharepoint_clients.sharepoint_client import SharePointClient


def find_worksheet(wb: Workbook, current_letter: str) -> Worksheet:
    """Finds worksheet in workbook corresponding to current letter"""
    if len(wb.worksheets) > 0:  # initialize and check
        ws = wb.worksheets[0]
    else:
        raise ValueError("No worksheets found in searched workbook")

    ws_found = False

    for sheet in wb.worksheets:  # get sheet corresponding to current_letter
        if sheet.title.strip() == current_letter:
            ws = sheet
            ws_found = True

    if not ws_found:
        raise ValueError(f"No sheet named {current_letter} found")

    return ws


def get_latest_company_name(wb: Workbook, current_letter: str) -> str:
    """
    This module reads the latest company that has been KYC-ed
    from the spreadsheet corresponding to current_letter.
    Note that worksheets are mutable inside a Workbook.
    """

    ws = find_worksheet(wb, current_letter)

    last_cell = get_last_filled_cell(ws, "B")

    if last_cell:
        logging.info("%s", last_cell.value)
        print(last_cell.value)
        return str(last_cell.value)

    logging.error("No cell found in worksheet %s", ws.title)
    raise ValueError(f"No cell found in worksheet '{ws.title}'")


def get_last_filled_cell(ws: Worksheet, column: Union[int, str]) -> Optional[Cell]:
    """
    Returns the last non-None cell in the given column.
    Union has the same effect as the pipe operator | but is compatible with Python < 3.10
    In fact, Optional[X] just means Union[X, None]
    """

    # Convert letter to number
    if isinstance(column, str):
        from openpyxl.utils import column_index_from_string

        col_idx = column_index_from_string(column)
    else:
        col_idx = column

    for row in range(ws.max_row, 0, -1):  # parameters are range(start, stop, step)
        cell = ws.cell(row=row, column=col_idx)
        if cell.value is not None:
            return cell
    return None


def get_top_left_cell(ws: Worksheet, cell: Cell | MergedCell) -> Cell:
    """Converts all merged cells in a worksheet to reflect its top left cell.
    Our script has no merged cells so it should be fine."""
    for merged_range in ws.merged_cells.ranges:
        if cell.coordinate in merged_range:
            min_col, min_row, max_col, max_row = range_boundaries(str(merged_range))
            if max_row:  # just here to satisfy type checker. ignore this block
                pass
            if max_col:
                pass
            return ws.cell(row=min_row, col=min_col)  # type: ignore
    return cell  # type: ignore[return-value]


def write_to_cell(ws: Worksheet, input_cell: Cell | MergedCell, write_value: str):
    """Writes a str value to a cell."""
    cell = get_top_left_cell(ws, input_cell)
    cell.value = write_value


def append_excel(
    wb: Workbook,
    current_letter: str,
    kah_list: List[Individual],
    current_company: "SharePointClient",
) -> Workbook:
    """Appends KAH information to the existing excel sheet"""

    ws = find_worksheet(wb, current_letter)
    last_cell = get_last_filled_cell(ws, "C")
    if last_cell is not None:
        company_cell = get_top_left_cell(ws, last_cell.offset(1, -1))
        company_cell.value = current_company.name
        for i, individual in enumerate(kah_list):
            if individual.role is not None:
                write_to_cell(ws, company_cell.offset(i, 1), individual.role)
            write_to_cell(ws, company_cell.offset(i, 2), individual.name)
            write_to_cell(ws, company_cell.offset(i, 3), individual.chinese_name)
            write_to_cell(ws, company_cell.offset(i, 4), individual.id_type)
            write_to_cell(ws, company_cell.offset(i, 5), individual.id_value)
            write_to_cell(ws, company_cell.offset(i, 6), individual.id_status)
            write_to_cell(ws, company_cell.offset(i, 7), str(individual.google))
            write_to_cell(ws, company_cell.offset(i, 8), str(individual.baidu))
            write_to_cell(ws, company_cell.offset(i, 12), str(individual.kyc_status))

    return wb
